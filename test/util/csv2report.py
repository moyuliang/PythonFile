# 根据禅道导出的csv更新测试报告

import csv, openpyxl, pandas, re
from openpyxl import Workbook


# csvfilename, excelfilename
def csv2report():
    # 读取csv文件
    with open("./../resource/input/交易平台-Bug.csv", encoding='gbk') as reader:
        content = list(csv.reader(reader))

    # 创建二维列表
    lists = [[] for i in range(len(content))]
    lists1 = [[] for i in range(len(content))]

    # 遍历
    for i in range(len(content)):
        for j in range(len(content[i])):
            #  将解决方案为“已解决”、“延期处理”、“”的bug编号、所属模块、严重程度（加上标题行）放入bug统计列表
            if j == 0 or j == 2 or j == 8:
                if i == 0 or content[i][25] == "已解决" or content[i][25] == "延期处理" or content[i][25] == "":
                    lists[i].append(content[i][j])

            # 将解决方案为“延期处理”、“”的bug编号、标题、解决方案（加上标题行）放入遗留bug列表
            if j == 0 or j == 6 or j == 25:
                if i == 0 or content[i][25] == "延期处理" or content[i][25] == "":
                    lists[i].append(content[i][j])

    # 去除空元素
    lists = [x for x in lists if x]
    lists1 = [x for x in lists1 if x]

    # 处理bug统计分析中的所属模块
    for i in range(1, len(lists)):
        # 形如：/(#0) ，/江苏(#629)，/四川 这类只有1个/的所属模块，统一处理为“通用”
        if lists[i][1].count("/") == 1 and "微信小程序" != re.search(r'''\/([\u4e00-\u9fa5]*)[0-9#()]*$''',
                                                                lists[i][1]).group(1):
            lists[i][1] = "通用"

        # /微信小程序

        # 形如：/江苏/后台($633) ，/江苏/前台/账户中心/售电报价管理(#687) 这类超过1个/的所属模块，均取第二个/后的文字为模块内容
        if lists[i][1].count("/") > 1:
            lists[i][1] = re.search(r'''\/[\u4e00-\u9fa5]*\/([\u4e00-\u9fa5]*)''', lists[i][1]).group(1)

    print(lists)

    # 创建Excel
    wb = Workbook()
    # ws=wb.create_sheet("bug统计分析")
    ws = wb.active
    ws.title = "bug统计分析"

    for row in lists:
        ws.append(row)

    ws1 = wb.create_sheet("遗留bug")

    for row in lists1:
        ws1.append(row)

    ws.save("./../resource/output/tmp.xlsx")


csv2report()
# str1 = "/四川/后台/信息管理(#365)"
# result = re.search(r'''(^\/([\u4e00-\u9fa5]*)[0-9#()]*$|\/([\u4e00-\u9fa5]*)\/)''', str1).group(3)
# result=re.findall(r'''(^\/([\u4e00-\u9fa5]*)[0-9#()]*$|\/([\u4e00-\u9fa5]*)\/)''', str1)
# print(result)
